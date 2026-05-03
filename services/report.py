"""
services/report.py — Geração de relatórios e exportação para Excel por usuário.
"""

import logging
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from services.storage import StorageService

logger = logging.getLogger(__name__)
storage = StorageService()


class ReportService:

    def resumo_mensal(self, user_id: int, periodo: str) -> dict[str, Any]:
        """Calcula totais de gastos, receitas e saldo para o usuário no período."""
        df = storage.carregar_dados(user_id)
        if df.empty:
            return self._vazio()

        df_periodo = df[df["data"].dt.strftime("%Y-%m") == periodo]
        if df_periodo.empty:
            return self._vazio()

        gastos = df_periodo[df_periodo["tipo"] == "gasto"]
        receitas = df_periodo[df_periodo["tipo"] == "receita"]

        total_gastos = round(gastos["valor"].sum(), 2)
        total_receitas = round(receitas["valor"].sum(), 2)

        por_categoria = {}
        if not gastos.empty:
            por_categoria = gastos.groupby("categoria")["valor"].sum().round(2).to_dict()

        return {
            "total_gastos": total_gastos,
            "total_receitas": total_receitas,
            "saldo": round(total_receitas - total_gastos, 2),
            "por_categoria": por_categoria,
        }

    @staticmethod
    def _vazio() -> dict:
        return {"total_gastos": 0.0, "total_receitas": 0.0, "saldo": 0.0, "por_categoria": {}}

    def exportar_xlsx(self, user_id: int, periodo: str, caminho: str) -> bool:
        """Gera planilha Excel com os dados do usuário no período."""
        df = storage.carregar_dados(user_id)
        if df.empty:
            return False

        df_periodo = df[df["data"].dt.strftime("%Y-%m") == periodo].copy()
        if df_periodo.empty:
            return False

        df_periodo["data"] = df_periodo["data"].dt.strftime("%d/%m/%Y %H:%M")
        df_display = df_periodo[["data", "tipo", "categoria", "valor"]].rename(columns={
            "data": "Data/Hora", "tipo": "Tipo",
            "categoria": "Categoria", "valor": "Valor (R$)",
        })

        resumo = self.resumo_mensal(user_id, periodo)
        resumo_rows = [
            ["Descrição", "Valor (R$)"],
            ["Total de Receitas", resumo["total_receitas"]],
            ["Total de Gastos", resumo["total_gastos"]],
            ["Saldo Final", resumo["saldo"]],
        ]
        if resumo["por_categoria"]:
            resumo_rows += [["", ""], ["Gastos por Categoria", ""]]
            for cat, val in sorted(resumo["por_categoria"].items(), key=lambda x: -x[1]):
                resumo_rows.append([cat.capitalize(), val])

        df_resumo = pd.DataFrame(resumo_rows[1:], columns=resumo_rows[0])

        with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
            df_display.to_excel(writer, sheet_name="Lançamentos", index=False)
            df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
            self._formatar(writer.sheets["Lançamentos"], valor_col=4)
            self._formatar(writer.sheets["Resumo"], valor_col=2, resumo=True)

        return True

    @staticmethod
    def _formatar(ws, valor_col: int, resumo: bool = False) -> None:
        cab_font = Font(bold=True, color="FFFFFF", size=11)
        cab_fill = PatternFill("solid", fgColor="1F4E79")
        borda = Border(**{s: Side(style="thin", color="BFBFBF")
                         for s in ("left", "right", "top", "bottom")})

        col_widths = {1: 22, 2: 14, 3: 20, 4: 14}
        for col, w in col_widths.items():
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = w

        for cell in ws[1]:
            cell.font = cab_font
            cell.fill = cab_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor="DDEEFF" if row_idx % 2 == 0 else "FFFFFF")
            for cell in row:
                cell.fill = fill
                cell.border = borda
                cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=valor_col).number_format = 'R$ #,##0.00'
            if not resumo:
                tipo_cell = ws.cell(row=row_idx, column=2)
                if tipo_cell.value == "gasto":
                    tipo_cell.font = Font(color="C00000", bold=True)
                elif tipo_cell.value == "receita":
                    tipo_cell.font = Font(color="375623", bold=True)
